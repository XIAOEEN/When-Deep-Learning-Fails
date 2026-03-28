"""
CKD Cat Data Loader
Loads and preprocesses all 6 Excel files into unified dataframes.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import warnings
warnings.filterwarnings('ignore')


# Treatment group cat IDs (confirmed from data: columns 4-6 in blood routine sheet)
TREATMENT_IDS = ['9711', '2793', '6424']
# Control group cat IDs (columns 7-12 in blood routine sheet)
CONTROL_IDS = ['2786', '2616', '9443', '9716', '9448', '9637']


def parse_date_from_sheet_name(sheet_name: str) -> Optional[datetime]:
    """Parse date from sheet name like '20250923' or '2025-09-23'"""
    formats = ['%Y%m%d', '%Y-%m-%d', '%Y%m%dT%H:%M:%S']
    s = str(sheet_name).strip()
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None


def load_blood_routine(data_dir) -> pd.DataFrame:
    """Load blood routine data from Excel."""
    data_dir = Path(data_dir)
    wb = openpyxl.load_workbook(data_dir / '血常规数据记录.xlsx', data_only=True)

    all_data = []
    for sn in wb.sheetnames:
        date = parse_date_from_sheet_name(sn)
        if date is None:
            continue
        ws = wb[sn]

        # Find the row with cat IDs (row after header row)
        header_row_idx = None
        id_row_idx = None
        cat_cols = []
        cat_ids = []

        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            row = list(row)
            # Header row: first cell is '项目' or '检测项目'
            if row[0] in ['项目', '检测项目']:
                header_row_idx = i
            # ID row: first 3 cells are None/empty, then contains cat IDs
            elif header_row_idx is not None and id_row_idx is None:
                # Check if this row has cat IDs in it
                potential_ids = [v for v in row[3:] if v is not None and str(v).strip()]
                if potential_ids:
                    id_row_idx = i
                    # Find columns with cat IDs
                    for j, v in enumerate(row):
                        if v is not None and str(v).strip():
                            try:
                                int(str(v).strip())
                                cat_cols.append(j)
                                cat_ids.append(str(v).strip())
                            except:
                                pass
                    break

        if header_row_idx is None or id_row_idx is None:
            continue

        # Read variables (rows after ID row)
        for row in rows[id_row_idx + 1:]:
            row = list(row)
            if row[0] and str(row[0]).strip() and row[0] not in ['项目', '检测项目']:
                var_name = str(row[0]).split('（')[0].split('（')[0].strip()
                for j, cat_id in zip(cat_cols, cat_ids):
                    if j < len(row):
                        val = row[j]
                        if val is not None and val != '':
                            try:
                                val = float(val)
                            except:
                                continue
                            all_data.append({
                                'date': date,
                                'cat_id': cat_id,
                                'variable': f'BR_{var_name}',
                                'value': val
                            })
    wb.close()

    if all_data:
        return pd.DataFrame(all_data)
    return pd.DataFrame(columns=['date', 'cat_id', 'variable', 'value'])


def load_blood_bio(data_dir) -> pd.DataFrame:
    """Load blood biochemistry data from Excel."""
    data_dir = Path(data_dir)
    wb = openpyxl.load_workbook(data_dir / '血生化数据记录.xlsx', data_only=True)

    all_data = []
    for sn in wb.sheetnames:
        date = parse_date_from_sheet_name(sn.split('初始')[0].strip() if '初始' in sn else sn)
        if date is None:
            continue
        ws = wb[sn]

        # Find header row
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row = list(row)
            if any(str(v) in TREATMENT_IDS + CONTROL_IDS for v in row[3:7] if v):
                header_row = i
                cat_cols = [j for j, v in enumerate(row) if v and str(v) in TREATMENT_IDS + CONTROL_IDS]
                cat_ids = [str(row[j]) for j in cat_cols]
                break

        if header_row is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0] and str(row[0]) not in ['项目名称', '项目全称']:
                var_name = str(row[0]).strip()
                for j, cat_id in zip(cat_cols, cat_ids):
                    val = row[j]
                    if val is not None and val != '':
                        try:
                            val = float(val)
                        except:
                            continue
                        all_data.append({
                            'date': date,
                            'cat_id': cat_id,
                            'variable': f'BB_{var_name}',
                            'value': val
                        })
    wb.close()

    if all_data:
        return pd.DataFrame(all_data)
    return pd.DataFrame(columns=['date', 'cat_id', 'variable', 'value'])


def load_urine_routine(data_dir) -> pd.DataFrame:
    """Load urine routine data from Excel."""
    data_dir = Path(data_dir)
    wb = openpyxl.load_workbook(data_dir / '尿常规数据记录.xlsx', data_only=True)

    all_data = []
    for sn in wb.sheetnames:
        date = parse_date_from_sheet_name(sn)
        if date is None:
            continue
        ws = wb[sn]

        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row = list(row)
            if any(str(v) in TREATMENT_IDS + CONTROL_IDS for v in row[2:6] if v):
                header_row = i
                cat_cols = [j for j, v in enumerate(row) if v and str(v) in TREATMENT_IDS + CONTROL_IDS]
                cat_ids = [str(row[j]) for j in cat_cols]
                break

        if header_row is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0] and str(row[0]) not in ['组别', '项目名称']:
                var_name = str(row[0]).split()[0].strip()
                for j, cat_id in zip(cat_cols, cat_ids):
                    val = row[j]
                    if val is not None and val != '':
                        try:
                            val = float(val)
                        except:
                            continue
                        all_data.append({
                            'date': date,
                            'cat_id': cat_id,
                            'variable': f'UR_{var_name}',
                            'value': val
                        })
    wb.close()

    if all_data:
        return pd.DataFrame(all_data)
    return pd.DataFrame(columns=['date', 'cat_id', 'variable', 'value'])


def load_weight_temp(data_dir) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Load weight and temperature records."""
    data_dir = Path(data_dir)
    wb = openpyxl.load_workbook(data_dir / '猫慢性肾病数据记录.xlsx', data_only=True)

    weight_data = []
    temp_data = []

    # Weight sheet
    if '体重记录' in wb.sheetnames:
        ws = wb['体重记录']
        # Find header row with dates
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row = list(row)
            if len(row) > 2 and row[0] == '组别':
                # Found header row
                date_cols = []
                dates = []
                for j, v in enumerate(row[2:], 2):
                    if isinstance(v, datetime):
                        dates.append(v)
                        date_cols.append(j)
                    elif isinstance(v, str) and len(v) == 8 and v.isdigit():
                        try:
                            d = datetime.strptime(v, '%Y%m%d')
                            dates.append(d)
                            date_cols.append(j)
                        except:
                            pass

                # Read weight data
                for row in ws.iter_rows(min_row=i + 1, values_only=True):
                    if row[1] and str(int(row[1])) in TREATMENT_IDS + CONTROL_IDS:
                        cat_id = str(int(row[1]))
                        group = 'treatment' if cat_id in TREATMENT_IDS else 'control'
                        for j, d in zip(date_cols, dates):
                            val = row[j]
                            if val is not None:
                                try:
                                    weight_data.append({
                                        'date': d,
                                        'cat_id': cat_id,
                                        'group': group,
                                        'variable': 'weight',
                                        'value': float(val)
                                    })
                                except:
                                    pass

    # Temperature sheet
    if '体温记录' in wb.sheetnames:
        ws = wb['体温记录']
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row = list(row)
            if len(row) > 2 and row[0] == '组别':
                date_cols = []
                dates = []
                for j, v in enumerate(row[2:], 2):
                    if isinstance(v, datetime):
                        dates.append(v)
                        date_cols.append(j)

                for row in ws.iter_rows(min_row=i + 1, values_only=True):
                    if row[1] and str(int(row[1])) in TREATMENT_IDS + CONTROL_IDS:
                        cat_id = str(int(row[1]))
                        for j, d in zip(date_cols, dates):
                            val = row[j]
                            if val is not None:
                                try:
                                    temp_data.append({
                                        'date': d,
                                        'cat_id': cat_id,
                                        'variable': 'temp',
                                        'value': float(val)
                                    })
                                except:
                                    pass
    wb.close()

    return (pd.DataFrame(weight_data) if weight_data else pd.DataFrame(),
            pd.DataFrame(temp_data) if temp_data else pd.DataFrame())


def load_fsaa(data_dir) -> pd.DataFrame:
    """Load fSAA (serum amyloid A) data."""
    data_dir = Path(data_dir)
    wb = openpyxl.load_workbook(data_dir / '血清fSAA.xlsx', data_only=True)

    all_data = []
    if 'fSAA' in wb.sheetnames:
        ws = wb['fSAA']

        # Get date row (row 0 has dates as column headers)
        header_row = list(ws.iter_rows(values_only=True))[0]  # row 1 (0-indexed = 0)

        # Find date columns
        date_cols = {}
        for j, v in enumerate(header_row):
            if isinstance(v, (int, float)) and v > 20250000:
                date_str = str(int(v))
                try:
                    d = datetime.strptime(date_str, '%Y%m%d')
                    date_cols[j] = d
                except:
                    pass
            elif isinstance(v, datetime):
                date_cols[j] = v

        # Read data rows (start from row 2, which is the first data row; row 1 is header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip rows where col1 (cat_id) is None or not a valid cat
            cat_id = str(int(row[1])) if row[1] is not None else None
            if cat_id and cat_id in TREATMENT_IDS + CONTROL_IDS:
                group = 'treatment' if cat_id in TREATMENT_IDS else 'control'
                for j, d in date_cols.items():
                    val = row[j]
                    if val is not None:
                        try:
                            all_data.append({
                                'date': d,
                                'cat_id': cat_id,
                                'group': group,
                                'variable': 'fSAA',
                                'value': float(val)
                            })
                        except:
                            pass

    wb.close()
    return pd.DataFrame(all_data) if all_data else pd.DataFrame()


def load_all_data(data_dir) -> Dict[str, pd.DataFrame]:
    """Load all data sources."""
    data_dir = Path(data_dir)
    print("Loading blood routine...")
    br = load_blood_routine(data_dir)
    print(f"  Blood routine: {len(br)} records, {br['cat_id'].nunique()} cats, {br['variable'].nunique()} variables")

    print("Loading blood biochemistry...")
    bb = load_blood_bio(data_dir)
    print(f"  Blood bio: {len(bb)} records, {bb['cat_id'].nunique()} cats, {bb['variable'].nunique()} variables")

    print("Loading urine routine...")
    ur = load_urine_routine(data_dir)
    print(f"  Urine routine: {len(ur)} records, {ur['cat_id'].nunique()} cats, {ur['variable'].nunique()} variables")

    print("Loading weight/temperature...")
    wt, tp = load_weight_temp(data_dir)
    print(f"  Weight: {len(wt)} records, {wt['cat_id'].nunique() if len(wt) else 0} cats")
    print(f"  Temperature: {len(tp)} records")

    print("Loading fSAA...")
    fsaa = load_fsaa(data_dir)
    print(f"  fSAA: {len(fsaa)} records, {fsaa['cat_id'].nunique() if len(fsaa) else 0} cats")

    return {
        'blood_routine': br,
        'blood_bio': bb,
        'urine_routine': ur,
        'weight': wt,
        'temp': tp,
        'fsaa': fsaa
    }


def create_unified_dataframe(data_dict: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Create unified dataframe with cat_id x date x variable matrix."""
    all_records = []

    for source, df in data_dict.items():
        if df.empty:
            continue
        for _, row in df.iterrows():
            record = {
                'date': row['date'],
                'cat_id': row['cat_id'],
                'variable': row['variable'],
                'value': row['value'],
                'source': source
            }
            if 'group' in row:
                record['group'] = row['group']
            all_records.append(record)

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame(all_records)
    if 'group' not in df.columns:
        df['group'] = df['cat_id'].apply(lambda x: 'treatment' if x in TREATMENT_IDS else 'control')

    return df


def get_pivot_table(df: pd.DataFrame, variables: List[str] = None) -> pd.DataFrame:
    """Create pivot table: rows=(cat_id, date), columns=variables, values=value."""
    if variables:
        df = df[df['variable'].isin(variables)]

    pivot = df.pivot_table(
        index=['cat_id', 'date'],
        columns='variable',
        values='value',
        aggfunc='first'
    ).reset_index()

    return pivot


if __name__ == '__main__':
    data_dir = Path('/Users/yangxiansen/Documents/CKD猫药物评价模型/Feline CKD 2')
    data = load_all_data(data_dir)
    unified = create_unified_dataframe(data)
    print(f"\nUnified data: {len(unified)} records")
    print(f"Cats: {unified['cat_id'].unique()}")
    print(f"Groups: {unified.groupby('cat_id')['group'].first().to_dict()}")
    print(f"Variables: {unified['variable'].unique()}")
